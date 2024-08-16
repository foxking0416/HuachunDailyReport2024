import os
import win32com.client
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.styles import PatternFill
from enum import Enum

current_dir = os.path.dirname(__file__)
parent_dir = os.path.dirname(current_dir)
# json_file_path = os.path.join(current_dir, 'ExternalData\\DailyReport.json')

image_path_no_data            = os.path.join(current_dir, 'ExternalData\\Image\\NoData.png') 
image_path_holiday            = os.path.join(current_dir, 'ExternalData\\Image\\Holiday.png') 
image_path_workday            = os.path.join(current_dir, 'ExternalData\\Image\\Workday.png') 
image_path_start_day          = os.path.join(current_dir, 'ExternalData\\Image\\StartDay.png') 
image_path_expect_finish_day  = os.path.join(current_dir, 'ExternalData\\Image\\ExpectFinishDay.png') 
image_path_real_finish_day    = os.path.join(current_dir, 'ExternalData\\Image\\RealFinishDay.png') 
image_path_sun_up             = os.path.join(current_dir, 'ExternalData\\Image\\Sun_Up.png')
image_path_sun_down           = os.path.join(current_dir, 'ExternalData\\Image\\Sun_Down.png')
image_path_rain_up            = os.path.join(current_dir, 'ExternalData\\Image\\Rain_Up.png')
image_path_rain_down          = os.path.join(current_dir, 'ExternalData\\Image\\Rain_Down.png') 
image_path_heavyrain_up       = os.path.join(current_dir, 'ExternalData\\Image\\HeavyRain_Up.png')
image_path_heavyrain_down     = os.path.join(current_dir, 'ExternalData\\Image\\HeavyRain_Down.png') 
image_path_typhoon_up         = os.path.join(current_dir, 'ExternalData\\Image\\Typhoon_Up.png')
image_path_typhoon_down       = os.path.join(current_dir, 'ExternalData\\Image\\Typhoon_Down.png')
image_path_hot_up             = os.path.join(current_dir, 'ExternalData\\Image\\Hot_Up.png')
image_path_hot_down           = os.path.join(current_dir, 'ExternalData\\Image\\Hot_Down.png')
image_path_muddy_up           = os.path.join(current_dir, 'ExternalData\\Image\\Muddy_Up.png')
image_path_muddy_down         = os.path.join(current_dir, 'ExternalData\\Image\\Muddy_Down.png')
image_path_weather_other_up   = os.path.join(current_dir, 'ExternalData\\Image\\WeatherOther_Up.png')
image_path_weather_other_down = os.path.join(current_dir, 'ExternalData\\Image\\WeatherOther_Down.png')

image_path_suspend_work_up    = os.path.join(current_dir, 'ExternalData\\Image\\SuspendWork_Up.png')
image_path_suspend_work_down  = os.path.join(current_dir, 'ExternalData\\Image\\SuspendWork_Down.png')
image_path_power_off_up       = os.path.join(current_dir, 'ExternalData\\Image\\PowerOff_Up.png')
image_path_power_off_down     = os.path.join(current_dir, 'ExternalData\\Image\\PowerOff_Down.png')
image_path_human_other_up     = os.path.join(current_dir, 'ExternalData\\Image\\HumanOther_Up.png')
image_path_human_other_down   = os.path.join(current_dir, 'ExternalData\\Image\\HumanOther_Down.png')


image_path_day_01 = os.path.join(current_dir, 'ExternalData\\Image\\Day_01.png')
image_path_day_02 = os.path.join(current_dir, 'ExternalData\\Image\\Day_02.png')
image_path_day_03 = os.path.join(current_dir, 'ExternalData\\Image\\Day_03.png')
image_path_day_04 = os.path.join(current_dir, 'ExternalData\\Image\\Day_04.png')
image_path_day_05 = os.path.join(current_dir, 'ExternalData\\Image\\Day_05.png')
image_path_day_06 = os.path.join(current_dir, 'ExternalData\\Image\\Day_06.png')
image_path_day_07 = os.path.join(current_dir, 'ExternalData\\Image\\Day_07.png')
image_path_day_08 = os.path.join(current_dir, 'ExternalData\\Image\\Day_08.png')
image_path_day_09 = os.path.join(current_dir, 'ExternalData\\Image\\Day_09.png')
image_path_day_10 = os.path.join(current_dir, 'ExternalData\\Image\\Day_10.png')
image_path_day_11 = os.path.join(current_dir, 'ExternalData\\Image\\Day_11.png')
image_path_day_12 = os.path.join(current_dir, 'ExternalData\\Image\\Day_12.png')
image_path_day_13 = os.path.join(current_dir, 'ExternalData\\Image\\Day_13.png')
image_path_day_14 = os.path.join(current_dir, 'ExternalData\\Image\\Day_14.png')
image_path_day_15 = os.path.join(current_dir, 'ExternalData\\Image\\Day_15.png')
image_path_day_16 = os.path.join(current_dir, 'ExternalData\\Image\\Day_16.png')
image_path_day_17 = os.path.join(current_dir, 'ExternalData\\Image\\Day_17.png')
image_path_day_18 = os.path.join(current_dir, 'ExternalData\\Image\\Day_18.png')
image_path_day_19 = os.path.join(current_dir, 'ExternalData\\Image\\Day_19.png')
image_path_day_20 = os.path.join(current_dir, 'ExternalData\\Image\\Day_20.png')
image_path_day_21 = os.path.join(current_dir, 'ExternalData\\Image\\Day_21.png')
image_path_day_22 = os.path.join(current_dir, 'ExternalData\\Image\\Day_22.png')
image_path_day_23 = os.path.join(current_dir, 'ExternalData\\Image\\Day_23.png')
image_path_day_24 = os.path.join(current_dir, 'ExternalData\\Image\\Day_24.png')
image_path_day_25 = os.path.join(current_dir, 'ExternalData\\Image\\Day_25.png')
image_path_day_26 = os.path.join(current_dir, 'ExternalData\\Image\\Day_26.png')
image_path_day_27 = os.path.join(current_dir, 'ExternalData\\Image\\Day_27.png')
image_path_day_28 = os.path.join(current_dir, 'ExternalData\\Image\\Day_28.png')
image_path_day_29 = os.path.join(current_dir, 'ExternalData\\Image\\Day_29.png')
image_path_day_30 = os.path.join(current_dir, 'ExternalData\\Image\\Day_30.png')
image_path_day_31 = os.path.join(current_dir, 'ExternalData\\Image\\Day_31.png')


c2e = cm_to_EMU
# Calculated number of cells width or height from cm into EMUs
cellh = lambda x: c2e((x * 49.77)/99)
cellw = lambda x: c2e((x * (18.65-1.71))/10)
col_offset = cellw(0.1) #60984
row_up_offset = cellh(0.25) #45245
row_down_offset = cellh(1) #180981
p2e = pixels_to_EMU
whole_size = XDRPositiveSize2D(p2e(30), p2e(30))
half_size = XDRPositiveSize2D(p2e(30), p2e(15))

fill_yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
fill_green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

class DailyReportType(Enum):
    TYPE_A = 0
    TYPE_B = 1
    TYPE_C = 2
    TYPE_D = 3
    TYPE_E = 4
    TYPE_F = 5
    TYPE_G = 6

def excel_to_pdf(excel_file, pdf_file):
    # 创建Excel应用程序对象
    excel = win32com.client.Dispatch("Excel.Application")
    # 打开Excel文件
    wb = excel.Workbooks.Open(excel_file)

    sheet_index = 0
    for sheet in wb.Sheets:

        serial_number = 1
        filename, extension = os.path.splitext(pdf_file)
        filename = filename + '_' + sheet.Name
        output_pdf = filename + '.pdf'
        while os.path.exists(output_pdf):
            # 如果文件已经存在，则添加流水号并重新检查
            output_pdf = f"{filename}_{serial_number}{extension}"
            serial_number += 1

        # 选择要保存为PDF的工作表
        ws = wb.Worksheets[sheet_index]

        # 将Excel文件保存为PDF
        ws.ExportAsFixedFormat(0, output_pdf)
        sheet_index += 1

    # 关闭Excel文件和应用程序
    wb.Close(False)
    excel.Quit()

def insert_image( worksheet, image_path, marker, size ):
    img = Image(image_path)
    img.anchor = OneCellAnchor(_from=marker, ext=size)
    worksheet.add_image(img)

def number_to_string(n):
    result = ''
    while n > 0:
        n -= 1
        result = chr(ord('A') + n % 26) + result
        n //= 26
    return result

